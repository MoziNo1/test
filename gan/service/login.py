import openpyxl
from selenium import webdriver
from gan.Page.basepage import PageCredit
from selenium.webdriver.common.action_chains import ActionChains
import time
from gan.element.getelement import Summary


class CreditLogin(PageCredit):

    name = input("请输入您的登录名：")

    def __init__(self, url):
        self.driver = webdriver.Chrome()
        self.url = url

    # 进行初始页面登录
    def home_login(self, login_name=name):
        # 创建对象获取元素
        self.getweb()
        self.driver.maximize_window()
        self.wait()
        self.input_text(str(login_name), '//*[@placeholder="请输入您的用户名"]')
        self.input_text(str(login_read('B')[0]), '//*[@placeholder="请输入您的密码"]')
        time.sleep(3)
        # 获取点击按钮,点击登录
        login_button = self.find_element('//*[@id="loginForm"]/div/div[7]/button')
        self.driver.execute_script("arguments[0].click();", login_button)
        # c.button_click('//*[@id="loginForm"]/div/div[7]/button')
        # self.find_element('//*[@id="loginForm"]/div/div[7]/button').click()
        # first_url = self.driver.current_url
        self.driver.implicitly_wait(10)
        time.sleep(1)
        try:
            if self.find_element(Summary.end_card_info):
                try:
                    self.button_click(Summary.end_card_info)
                except Exception as e:
                    pass
        except Exception as e:
            pass
        return login_name


def login_read(column, file_path='../../146/data/146_login_name.xlsx'):
    excel = openpyxl.open(file_path)
    worksheet = excel.worksheets[0]

    row = worksheet.max_row
    # column = worksheet.max_column

    login_list = []
    for rows in range(row + 1)[2:]:
        position = column + str(rows)
        login_name = worksheet[position].value
        login_list.append(login_name)
    return login_list


# 读取登录名以及用户名
#  发起人基本信息填写完毕后，将改登录名写入登录名的excel表中

def login_write(value, file_name):
    excel = openpyxl.load_workbook(file_name)
    worksheet = excel.active
    # print(worksheet['C2'].value)
    worksheet['C2'] = value
    # print(worksheet['C2'].value)
    excel.save(filename=file_name)
    excel.close()

    quota_data_list = login_read('B', '../../146/data/quota.xlsx')
    # print(quota_data_list)

def login_write_sql(value1, value2, file_name):
    excel = openpyxl.load_workbook(file_name)
    worksheet = excel.active
    # print(worksheet['C2'].value)
    worksheet['B2'] = value1
    worksheet['B4'] = value2
    # print(worksheet['C2'].value)
    excel.save(filename=file_name)
    excel.close()


def login_full_write(row, clomun, value, file_name):
    excel = openpyxl.load_workbook(file_name)
    worksheet = excel.active
    worksheet[f"{row}{clomun}"] = value
    excel.save(filename=file_name)
    excel.close()