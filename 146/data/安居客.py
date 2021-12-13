# -*- coding: utf-8 -*-
# @Time    :
# @Author  :
# @Email   :
# @File    :
import requests
import re
from bs4 import BeautifulSoup
import csv
import time
import threading
from lxml import etree
from selenium import webdriver
from openpyxl import Workbook

num0 = 1  # 用来计数
baseurl = 'https://www.anjuke.com/sy-city.html'

wb = Workbook()
ws = wb.active
ws.title = '安居客'
ws.cell(row=1, column=1).value = '城市链接'
ws.cell(row=1, column=2).value = '城市名称'

def gethtml():
    chromedriver = "chromedriver.exe"
    browser = webdriver.Chrome(chromedriver)
    browser.get(baseurl)
    time.sleep(5)

    #让页面滚动到下面,window.scrollBy(0, scrollStep),ScrollStep ：间歇滚动间距
    js = 'window.scrollBy(0,3000)'
    browser.execute_script(js)
    js = 'window.scrollBy(0,5000)'
    browser.execute_script(js)
    html = browser.page_source
    return html

def parseHotBook(html):
    # print(html)
    regAuthor = r'.*?<a href="(.*?)</a>'
    reg_author = re.compile(regAuthor)
    authorother = re.findall(reg_author, html)

    global num0

    for info in authorother:
        verinfo = info.split('">')
        print(verinfo[0],verinfo[1].replace('class="hot',''))

        num0 = num0 + 1
        name = verinfo[0]
        link = verinfo[1].replace('class="hot','')
        ws.cell(row=num0, column=1).value = name
        ws.cell(row=num0, column=2).value = link
    wb.save('安居客2' + '.xlsx')
    print('爬取成功')

if __name__=='__main__':
    html = gethtml()
    parseHotBook(html)
